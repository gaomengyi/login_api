# encoding=utf-8
from openpyxl import load_workbook
from test_case.get_data import GetData
from tool.get_config import *
from tool.project_path import *

"""
专门读取excel的值
"""


class GetExcel:
    def __init__(self, filename, sheet):
        self.filename = filename
        self.sheet = sheet

    # 第一种方法
    # def get_heard(self):
    #     get_file = load_workbook(self.filename)  # 打开文件
    #     get_sheet = get_file[self.sheet]  # 打开sheet页
    #     heard = []  # 创建一个空列表
    #     for i in range(1, get_sheet.max_column + 1):  # for循环第一列到最后一列
    #         heard.append(get_sheet.cell(1, i).value)  # for循环遍历取第一行的每一个列值
    #     return heard
    #
    # def get_data(self, button="all"):
    #     # 读取到的ini文件中的button值 button取值是配置文件中数据
    #     # button = ReadConfig.read_config(caseini_path,'MODE','button')
    #     get_file = load_workbook(self.filename)  # 打开文件
    #     get_sheet = get_file[self.sheet]  # 打开sheet页
    #     rownum = get_sheet.max_row  # 获得最大行值
    #     colnum = get_sheet.max_column  # 获得最大列值
    #     test_data = []
    #     header = self.get_heard()  # 获取到heard的值
    #     for i in range(2, rownum + 1):  # 从第二行开始循环到最大行值
    #         # 放在里面每次循环都会新建一个字典
    #         sub_data = {}  # 每一列都是字典格式 重要：这个一定要写在行内循环中，不然一直循环最后一行
    #         for j in range(1, colnum + 1):  # 从第一行开始循环到最大列值
    #             # 当i=1时，j=1,2,3,4,5....
    #             sub_data[header[j - 1]] = get_sheet.cell(i, j).value  # 相当于sub_data['case']=1,sub_data['tile']=‘正常登录’
    #         test_data.append(sub_data)  # 一定要在第一个for行循环下，不然每行数据重复读取多次
    #         # print(test_data)
    #     # 根据button判断用例是否全部执行
    #     if button == "all":
    #         final_data = test_data  # 如果button==all tets_data就是最终的测试数据
    #     else:
    #         final_data = []  # 如果不是，创建一个空字典，给final_data数据加进去
    #         for item in test_data:
    #             # 读取配置文件中数据都是str类型 需要转换
    #             if item['case'] in eval(button):  # 如果case的值在button中 给循环的数据加入到final_data中
    #                 final_data.append(item)
    #     print(final_data)
    #     return final_data

    # 第二种方法
    def get_data(self):
        # tel = "15712932600"
        # tel_1 = "admin"
        get_file = load_workbook(self.filename)  # 打开文件
        get_sheet = get_file[self.sheet]  # 打开sheet页
        rownum = get_sheet.max_row  # 获得最大行值
        tel = getattr(GetData, "tel")
        tel_1 = getattr(GetData, "tel_1")
        test_data = []
        for i in range(2, rownum + 1):
            sub_data = {}
            sub_data['case'] = get_sheet.cell(i, 1).value
            sub_data['modul'] = get_sheet.cell(i, 2).value
            sub_data['title'] = get_sheet.cell(i, 3).value
            sub_data['method'] = get_sheet.cell(i, 4).value
            sub_data['url'] = get_sheet.cell(i, 5).value
            # sub_data['data'] = get_sheet.cell(i,6).value
            if get_sheet.cell(i, 6).value.find("${tel}") != -1:  # 如果在第6列找到${tel}这个值 -1表示没有找到
                # 利用反射：如果在第6列能找到${tel}则进行下面的反射操作
                sub_data['data'] = get_sheet.cell(i, 6).value.replace('${tel}', str(tel))
                print(str(getattr(GetData, 'tel')))
            elif get_sheet.cell(i, 6).value.find("${tel_1") != -1:  # 如果在第6列找到${tel_1}这个值
                sub_data["data"] = get_sheet.cell(i, 6).value.replace('${tel_1}', str(tel_1))
            else:
                sub_data['data'] = get_sheet.cell(i, 6).value
            sub_data['expect'] = get_sheet.cell(i, 7).value
            test_data.append(sub_data)
        print(test_data)
        return test_data

    # 写一个方法：专门存储返回数据
    def data_write(self, row, col, test_result):
        file = load_workbook(self.filename)  # 打开文件
        sheet = file[self.sheet]  # 打开sheet页
        sheet.cell(row, col).value = test_result  # 定位到i行0弟8列写入test_result值
        # sheet.cell(row, col).value = result  # 定位到i行0弟9列写入result值
        file.save(self.filename)  # 保存文件


if __name__ == '__main__':
    res = GetExcel(testcase_path, 'login')
    res.get_data()
